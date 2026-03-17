"""
국민연금 입퇴사 분석 웹앱 (Streamlit)
- 회사명 검색 → 목록에서 선택 → 기간 설정 → 엑셀 다운로드
"""
import streamlit as st
import json, time, urllib.request, urllib.parse, io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

SERVICE_KEY = "3717e0d8ac3175e12853b60f074407885bee6b5987b6d7e4be02102004cb167e"
BASE = "https://api.odcloud.kr/api/15083277/v1"

ENDPOINTS = {
    "2015-12": "uddi:b9bf303e-a60e-4a49-a517-99797889484e",
    "2016-01": "uddi:37493930-4de8-4f00-8743-aeb803dcd795",
    "2016-02": "uddi:6ee79a68-53e3-4d63-ac62-a7979952e89b",
    "2016-03": "uddi:f2bba05e-c160-412f-800e-edc2327ab0f0",
    "2016-04": "uddi:42e4a970-2989-42bc-b357-94c31990d534",
    "2016-05": "uddi:735caf7c-2cd2-4570-a480-b4d9b57e1902",
    "2016-06": "uddi:9c5a563f-882a-4309-a05e-c7f6bc346bf5",
    "2016-07": "uddi:c11a6a3d-706c-422e-bca8-dca8bd93df8e",
    "2016-08": "uddi:73bd88e6-16e9-4803-ba58-a03604d52f74",
    "2016-09": "uddi:f430c5ab-3f7f-44f9-97f4-c327f0df7934",
    "2016-10": "uddi:3625c650-d32b-4ccf-be60-d3f0fcf962e6",
    "2016-11": "uddi:21e63777-8300-4e65-864e-b406bf408ebd",
    "2016-12": "uddi:dd4d69f9-2937-4785-83e0-b5fefdc361df",
    "2017-01": "uddi:c43047e8-9aea-42a3-ae44-c0e6d365f75e",
    "2017-02": "uddi:720f5972-ad02-49f8-b3a7-e26356f7c501",
    "2017-03": "uddi:a2ed5062-cc7f-4474-9e87-bcc0c9458ae9",
    "2017-04": "uddi:a201b487-f425-4f27-9378-f26cb9503ce6",
    "2017-05": "uddi:ad211696-f7b8-440b-b1d2-c14d28528bb7",
    "2017-06": "uddi:72aa6fc6-4d87-4156-bc91-e93289cdb0d1",
    "2017-07": "uddi:bf28351f-a1d6-4553-bb56-b753aef41fc2",
    "2017-08": "uddi:9fbeeafc-4bc8-4d42-8ac2-c2a1e090935b",
    "2017-09": "uddi:3ee261dc-747f-49dd-a157-f0bb17bc48e5",
    "2017-10": "uddi:0c963b82-5230-4aa3-9241-d157cb75b567",
    "2017-11": "uddi:a05d66c1-8056-4ed8-8197-b5bdd3b722f6",
    "2017-12": "uddi:f616dc6d-eb7a-45b1-a205-f91cdab72e40",
    "2018-01": "uddi:643c5346-a1cc-48e9-9217-d37c5de93c02",
    "2018-02": "uddi:49de868a-9fb3-4e6c-8835-f7718f9b446b",
    "2018-03": "uddi:2828565f-6cd9-4f9e-8326-91c394d451a3",
    "2018-04": "uddi:f3cdfbf4-12d8-40e0-9d97-bc7489e8820c",
    "2018-05": "uddi:c87d8aa5-cf12-40f9-8517-f9b5ab4d792c",
    "2018-06": "uddi:a9cb5774-238c-40f1-9d3a-badc3d3864d0",
    "2018-07": "uddi:202031a2-24eb-44b8-8c43-88f2de2b2d48",
    "2018-08": "uddi:658159ed-ce5d-4c09-9d7b-a1b2b7ab8089",
    "2018-09": "uddi:8b52fe84-5ed7-4232-a468-ca42091447a0",
    "2018-10": "uddi:30f90a31-09c9-41b6-9828-d05e2939db9f",
    "2018-11": "uddi:cf1753c5-0080-4c6d-b68e-cbeafa31f371",
    "2018-12": "uddi:088e933f-b16a-4002-9122-d4fb0c0f9832",
    "2019-01": "uddi:7f7319d8-df12-489c-9f1f-d46b987b05e9",
    "2019-02": "uddi:c310c1b7-4791-4bae-aea0-dfb81ce5f4cd",
    "2019-03": "uddi:2a2c6ef7-e956-464b-bc33-fcf33e9c2d0e",
    "2019-04": "uddi:8926074b-d33c-40fa-aa06-9e145b63c22d",
    "2019-05": "uddi:cb7aff6d-3e8b-4a0b-b5eb-8e055ef83d56_201908061631",
    "2019-06": "uddi:8a521a17-9426-4a43-ad4c-baecadf372da_201908061633",
    "2019-07": "uddi:c6bf89c2-8c0b-4c8e-8698-b2cd9dc31d1f_201908061635",
    "2019-08": "uddi:c6704217-b37b-48bf-a4a9-eeadf331688b_201908281108",
    "2019-09": "uddi:41296667-bc83-454f-bcfa-fd00ebacdab4",
    "2019-10": "uddi:1b4e4b0c-6f6c-4b43-8080-8b4c3746ddec",
    "2019-11": "uddi:ecd8b423-aa81-4c08-b1fe-b9037cd66a83",
    "2019-12": "uddi:02ad33d5-f387-4446-8884-a96bbf321a55",
    "2020-01": "uddi:32a88167-3191-47a5-8c58-e74f96393fb0",
    "2020-02": "uddi:3ede1edb-f56f-4248-af5c-d4b031333439",
    "2020-03": "uddi:fef9cdc9-8a2e-42bb-85af-82aee48119eb",
    "2020-04": "uddi:c56acab8-dcf0-40e6-9bca-d8260ff714a6",
    "2020-05": "uddi:79803eb8-0b53-423b-bf0d-e37a4dad398d",
    "2020-06": "uddi:088d6341-ebdb-473f-bdd7-b2e229404db8",
    "2020-07": "uddi:618aa96f-14c3-443b-bfa2-893f5902b8ea",
    "2020-08": "uddi:c38428d1-8408-4c4e-a814-f81651f1b8b6",
    "2020-09": "uddi:14afa7a7-becf-4e74-a440-c32e5c4ad19f",
    "2020-10": "uddi:d7b668da-60c3-4f4a-bcf9-9166dc5bc49e",
    "2020-11": "uddi:7eecccd0-c268-4c77-a4ae-a3c673ffa682",
    "2020-12": "uddi:7ff17f03-3072-4f20-b265-aa60ee041401",
    "2021-01": "uddi:b2a76201-9cf7-458d-a6e0-c2194e6ee40f",
    "2021-02": "uddi:d066a0fa-34e1-4149-9980-f5c9f7f9e2e3",
    "2021-03": "uddi:995de2fb-5dc2-4a3f-b0c7-c95deb374224",
    "2021-04": "uddi:f1d30736-6610-4a2d-a830-850c7da466e6",
    "2021-05": "uddi:0beb7252-2d35-4b74-af04-c77eed7ca44b",
    "2021-06": "uddi:d4781c30-5b0a-470b-a2b5-fd3e9cb20606",
    "2021-07": "uddi:9c7be522-8efa-427b-892f-ce95568d8779",
    "2021-08": "uddi:d54b642a-8db0-4359-a7f2-e8869b00913c",
    "2021-09": "uddi:3ee3626b-e28c-4f26-bba8-ff1b3f7fac0e",
    "2021-10": "uddi:dd18a482-3649-44b7-9fee-ed50b0aaaf9d",
    "2021-11": "uddi:2e4217d6-7c19-4e10-bcb6-a71e83a733bd",
    "2021-12": "uddi:4db30f80-d2a1-4a90-a89f-80a118ae4b85",
    "2022-01": "uddi:aa411324-4620-4b1f-be57-8e04b1e6c87c",
    "2022-01b": "uddi:9bdcbdbb-5402-4028-9a90-eaef6492e208",
    "2022-02": "uddi:6dea9362-4865-4ede-8b7f-98683b206668",
    "2022-03": "uddi:2ead034f-65e0-4d5d-b72e-f8dedba041a3",
    "2022-04": "uddi:23712302-15c4-4034-96b1-8ec80a415e44",
    "2022-05": "uddi:d7e2de87-da03-4ec4-9741-ef4208ce393c",
    "2022-06": "uddi:da7e3a30-3401-4232-87b1-b279d0d97088",
    "2022-07": "uddi:814e0e45-ab16-401d-92d6-bfc51260eba5",
    "2022-08": "uddi:d2683037-b144-466f-b287-90f79a4bd8b2",
    "2022-09": "uddi:e9b54ba9-04e7-4730-9def-db6f05bf1925",
    "2022-10": "uddi:ffe2743c-f90d-4644-bfeb-c770d319bec5",
    "2022-12": "uddi:9a6668fe-9df7-4118-9e1c-9fec68de7c03",
    "2023-01": "uddi:e80b5227-bbd2-4620-ae95-b82064f54da0",
    "2023-02": "uddi:86233d7b-8f57-4ed7-aab8-b59464141971",
    "2023-03": "uddi:e825a3c5-86c1-48b0-9450-a898731ec064",
    "2023-04": "uddi:e2be1bab-b4b5-4073-97dd-8a8487ecb487",
    "2023-05": "uddi:63837df0-61d0-4aa8-8890-ff3720b255f3",
    "2023-06": "uddi:84b05020-c26e-4a57-9a15-c60554764534",
    "2023-07": "uddi:b32e56a7-17cd-4ebf-8833-ac01662800da",
    "2023-08": "uddi:8617a7f1-3665-48d0-8b71-922cc1bdca07",
    "2023-09": "uddi:8b59078d-9a70-489f-98d8-ba27d7e573c5",
    "2023-10": "uddi:f42b3ec1-4fe9-4338-a282-91f61dc7f288",
    "2023-11": "uddi:2c0534ef-64f5-4f65-a627-b7ac918390dc",
    "2023-12": "uddi:00ade2e8-46b6-436d-9287-ae03d5a63a6f",
    "2024-01": "uddi:d3ffb517-deee-4b1e-8e9e-5e3e3fae54c7",
    "2024-01b": "uddi:5cdf7e9e-dc5d-4369-96d3-d22744d8d10d",
    "2024-02": "uddi:c70b85ac-0146-41a9-8f4a-d2acafaa3c92",
    "2024-03": "uddi:67ccdcc5-727f-408d-802a-dce95772acb8",
    "2024-03b": "uddi:f2d5995e-5b47-4476-9f04-c8b2519735a3",
    "2024-04": "uddi:ccc6764a-b232-494e-a453-d21b929878f8",
    "2024-05": "uddi:5ae3f030-6646-4239-b0f5-8e1b7b284007",
    "2024-06": "uddi:fbc9aff6-7496-4c14-bc49-adfefb93557d",
    "2024-07": "uddi:3f8e431e-efcf-4d25-b6f6-cef316722b84",
    "2024-08": "uddi:ae2d6a33-e33b-4312-a902-c9a8c22d9ab0",
    "2024-09": "uddi:a1d51e9d-f55a-4f94-a06c-ef98691479fd",
    "2024-10": "uddi:f6873590-8c0d-4328-af24-8b24f81deb8b",
    "2024-11": "uddi:819641d2-a9ba-498a-a689-db8ad2c9800a",
    "2024-12": "uddi:3a89a14e-7230-467a-bf07-9ca33d06812d",
    "2025-01": "uddi:45ba8ffb-ab8c-44da-abd6-b10ec30821cd",
    "2025-02": "uddi:6d064493-1c29-4ddb-9bc5-be98e40a1e57",
    "2025-03": "uddi:45b0b01c-16bd-4621-ad04-fdaeb400b4f6",
    "2025-04": "uddi:6ec70fba-037c-4e20-8d47-88e26912b4e2",
    "2025-05": "uddi:58d465f8-71bf-4378-b4e8-e4b265e805da",
    "2025-06": "uddi:7c8ebb8e-baf4-49a0-a281-aa483c3158b8",
    "2025-07": "uddi:20ddf65d-51d8-421f-8ee5-b64f05554151",
    "2025-08": "uddi:14c0beb5-b153-4b03-892b-8d30a7600de1",
    "2025-09": "uddi:466a4aef-5a2d-4b2b-a3d9-8a6c11b81d23",
    "2025-10": "uddi:f9787983-d48a-4c94-b7b6-c805a5be3cca",
    "2025-11": "uddi:06b329ca-54a4-47f8-8c8f-8268d61c7d7c",
    "2025-12": "uddi:10a6e7bd-a2ee-4ee1-967c-bb9a6aea89a9",
    "2026-01": "uddi:74d9aa39-bc2c-4124-9cff-ec389dbf51e3",
}

SEARCH_UUIDS = [
    ("2025-01", "uddi:45ba8ffb-ab8c-44da-abd6-b10ec30821cd"),
    ("2024-12", "uddi:3a89a14e-7230-467a-bf07-9ca33d06812d"),
    ("2024-06", "uddi:fbc9aff6-7496-4c14-bc49-adfefb93557d"),
]


# ──────────────────── API 함수 ────────────────────

def get_field(r, *candidates, default=0):
    for key in candidates:
        for rk in r:
            if key in rk and r[rk] is not None:
                val = r[rk]
                if isinstance(val, (int, float)):
                    return val
                if isinstance(val, str):
                    val = val.strip().replace(',', '')
                    if val == '' or val == '-':
                        continue
                    try:
                        return int(val)
                    except ValueError:
                        try:
                            return float(val)
                        except ValueError:
                            return val
                return val
    return default


def parse_record(r, endpoint_ym):
    members = get_field(r, '가입자수')
    new = get_field(r, '신규취득자수', 'col-20')
    lost = get_field(r, '상실가입자수', 'col-21')
    amount = get_field(r, '당월고지금액', 'col-19')
    data_ym = get_field(r, '자료생성년월', default='')
    actual_ym = data_ym if data_ym and len(str(data_ym)) >= 7 else endpoint_ym
    return {
        'ym': str(actual_ym)[:7],
        'name': r.get('사업장명', ''),
        'members': members if isinstance(members, (int, float)) else 0,
        'new': new if isinstance(new, (int, float)) else 0,
        'lost': lost if isinstance(lost, (int, float)) else 0,
        'address': r.get('사업장도로명상세주소') or r.get('사업장지번상세주소', ''),
        'biz_type': r.get('사업장업종코드명', ''),
        'monthly_amount': amount,
        'status': get_field(r, '사업장가입상태코드', default=0),
        'applied_date': r.get('적용일자', ''),
        'zipcode': r.get('우편번호', ''),
        'data_ym': str(data_ym),
    }


def api_fetch(uuid, search_name, page=1, per_page=100):
    cond_key = urllib.parse.quote("cond[사업장명::LIKE]")
    url = (f"{BASE}/{uuid}?page={page}&perPage={per_page}"
           f"&serviceKey={SERVICE_KEY}"
           f"&{cond_key}={urllib.parse.quote(search_name)}")
    try:
        req = urllib.request.Request(url)
        with urllib.request.urlopen(req, timeout=15) as resp:
            result = json.loads(resp.read().decode('utf-8'))
            return result.get('data', [])
    except Exception:
        return []


def search_companies(keyword):
    found = {}
    for ym, uuid in SEARCH_UUIDS:
        rows = api_fetch(uuid, keyword)
        for r in rows:
            name = r.get('사업장명', '').strip()
            if not name:
                continue
            members = get_field(r, '가입자수')
            biz = r.get('사업장업종코드명', '')
            addr = r.get('사업장도로명상세주소') or r.get('사업장지번상세주소', '')
            if name not in found or (isinstance(members, (int, float)) and members > 0):
                found[name] = {
                    'name': name,
                    'members': members if isinstance(members, (int, float)) else 0,
                    'biz_type': biz,
                    'address': addr,
                }
        time.sleep(0.1)
    return sorted(found.values(), key=lambda x: -x['members'])


# ──────────────────── 분석 & 엑셀 ────────────────────

def run_analysis(exact_name, start_year, progress_bar, status_text):
    filtered = {}
    for ym, uuid in ENDPOINTS.items():
        base_ym = ym.rstrip('b')
        yr = int(base_ym[:4])
        if yr >= start_year:
            filtered[ym] = uuid

    total = len(filtered)
    all_records = []
    seen_ym = set()

    for i, (ym, uuid) in enumerate(sorted(filtered.items())):
        rows = api_fetch(uuid, exact_name)
        for r in rows:
            rec = parse_record(r, ym)
            if rec['name'] == exact_name and rec['members'] > 0 and rec['ym'] not in seen_ym:
                all_records.append(rec)
                seen_ym.add(rec['ym'])
        progress_bar.progress((i + 1) / total, text=f"{ym} 조회 중... ({i+1}/{total})")
        time.sleep(0.12)

    if not all_records:
        status_text.error(f"'{exact_name}' 데이터가 없습니다.")
        return None

    # 월별 집계
    monthly = {}
    for r in all_records:
        ym = r['ym']
        if ym not in monthly:
            monthly[ym] = {'members': 0, 'new': 0, 'lost': 0}
        monthly[ym]['members'] += r['members']
        monthly[ym]['new'] += r['new']
        monthly[ym]['lost'] += r['lost']

    # 연속 월 보간
    all_yms = sorted(monthly.keys())
    first_ym, last_ym = all_yms[0], all_yms[-1]
    fy, fm = int(first_ym[:4]), int(first_ym[5:7])
    ly, lm = int(last_ym[:4]), int(last_ym[5:7])
    full_yms = []
    cy, cm = fy, fm
    while (cy, cm) <= (ly, lm):
        full_yms.append(f"{cy}-{cm:02d}")
        cm += 1
        if cm > 12:
            cm = 1; cy += 1

    for ym in full_yms:
        if ym not in monthly:
            idx = full_yms.index(ym)
            prev = nxt = None
            for j in range(idx - 1, -1, -1):
                if full_yms[j] in monthly:
                    prev = monthly[full_yms[j]]; break
            for j in range(idx + 1, len(full_yms)):
                if full_yms[j] in monthly:
                    nxt = monthly[full_yms[j]]; break
            if prev and nxt:
                members = round((prev['members'] + nxt['members']) / 2)
            elif prev:
                members = prev['members']
            else:
                members = nxt['members'] if nxt else 0
            monthly[ym] = {'members': members, 'new': 0, 'lost': 0, 'interpolated': True}

    # 연도별 집계
    year_data = {}
    for ym, d in monthly.items():
        yr = ym[:4]
        if yr not in year_data:
            year_data[yr] = {'sum_members': 0, 'count': 0, 'total_new': 0, 'total_lost': 0}
        year_data[yr]['sum_members'] += d['members']
        year_data[yr]['count'] += 1
        year_data[yr]['total_new'] += d['new']
        year_data[yr]['total_lost'] += d['lost']

    return make_excel(exact_name, all_records, monthly, year_data)


def make_excel(company_name, all_records, monthly, year_data):
    wb = openpyxl.Workbook()
    hdr_font = Font(bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    yr_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    interp_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    interp_font = Font(italic=True, color="999999")
    pct = '0.0%'

    def write_hdr(ws, r, cols):
        for c, h in enumerate(cols, 1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin

    def auto_width(ws):
        for col in ws.columns:
            ml = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 4, 30)

    # Sheet1: 원데이터
    ws1 = wb.active; ws1.title = "원데이터(Raw)"
    cols1 = ['기준월', '사업장명', '가입자수(현원)', '신규취득(입사)', '상실(퇴사)',
             '당월고지금액', '업종', '주소', '우편번호', '적용일자', '상태코드', '자료생성월']
    write_hdr(ws1, 1, cols1)
    for i, r in enumerate(sorted(all_records, key=lambda x: x['ym']), 2):
        vals = [r['ym'], r['name'], r['members'], r['new'], r['lost'],
                r['monthly_amount'], r['biz_type'], r['address'],
                r['zipcode'], r['applied_date'], r['status'], r['data_ym']]
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row=i, column=c, value=v)
            cell.border = thin
            if c == 6: cell.number_format = '#,##0'
    auto_width(ws1)

    # Sheet2: 월별현황
    ws2 = wb.create_sheet("월별현황")
    cols2 = ['기준월', '현원', '입사', '퇴사', '순증감', '입사율', '퇴사율', '비고']
    write_hdr(ws2, 1, cols2)
    row = 2; prev_year = None
    for ym in sorted(monthly.keys()):
        year = ym[:4]
        if prev_year and year != prev_year and prev_year in year_data:
            yd = year_data[prev_year]
            am = yd['sum_members'] / yd['count']
            vals = [f"{prev_year}년 소계", round(am), yd['total_new'], yd['total_lost'],
                    yd['total_new'] - yd['total_lost'],
                    yd['total_new'] / am if am else 0, yd['total_lost'] / am if am else 0, '']
            for c, v in enumerate(vals, 1):
                cell = ws2.cell(row=row, column=c, value=v)
                cell.fill = yr_fill; cell.font = Font(bold=True); cell.border = thin
                if c in (6, 7): cell.number_format = pct
            row += 1
        d = monthly[ym]
        is_interp = d.get('interpolated', False)
        net = d['new'] - d['lost']
        vals = [ym, d['members'], d['new'], d['lost'], net,
                d['new'] / d['members'] if d['members'] else 0,
                d['lost'] / d['members'] if d['members'] else 0,
                '추정(API 누락)' if is_interp else '']
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=c, value=v)
            cell.border = thin
            if c in (6, 7): cell.number_format = pct
            if is_interp:
                cell.fill = interp_fill
                if c == 8: cell.font = interp_font
        row += 1; prev_year = year
    if prev_year and prev_year in year_data:
        yd = year_data[prev_year]
        am = yd['sum_members'] / yd['count']
        vals = [f"{prev_year}년 소계", round(am), yd['total_new'], yd['total_lost'],
                yd['total_new'] - yd['total_lost'],
                yd['total_new'] / am if am else 0, yd['total_lost'] / am if am else 0, '']
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=c, value=v)
            cell.fill = yr_fill; cell.font = Font(bold=True); cell.border = thin
            if c in (6, 7): cell.number_format = pct
    auto_width(ws2)

    # Sheet3: 연도별요약
    ws3 = wb.create_sheet("연도별요약")
    cols3 = ['연도', '평균현원', '연간입사', '연간퇴사', '순증감', '입사율', '퇴사율',
             '데이터월수', '기초인원', '기말인원']
    write_hdr(ws3, 1, cols3)
    row = 2
    for yr in sorted(year_data.keys()):
        yd = year_data[yr]
        am = yd['sum_members'] / yd['count']
        ym_list = sorted([m for m in monthly if m.startswith(yr)])
        vals = [yr, round(am), yd['total_new'], yd['total_lost'],
                yd['total_new'] - yd['total_lost'],
                yd['total_new'] / am if am else 0, yd['total_lost'] / am if am else 0,
                yd['count'], monthly[ym_list[0]]['members'], monthly[ym_list[-1]]['members']]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row=row, column=c, value=v)
            cell.border = thin
            if c in (6, 7): cell.number_format = pct
        row += 1
    auto_width(ws3)

    # Sheet4: 분기별
    ws4 = wb.create_sheet("분기별현황")
    cols4 = ['분기', '기초인원', '입사', '퇴사', '기말인원', '순증감', '퇴사율']
    write_hdr(ws4, 1, cols4)
    row = 2
    for yr in sorted(year_data.keys()):
        for q in range(1, 5):
            ms = (q - 1) * 3 + 1
            qm = [f"{yr}-{m:02d}" for m in range(ms, ms + 3) if f"{yr}-{m:02d}" in monthly]
            if not qm: continue
            qn = sum(monthly[m]['new'] for m in qm)
            ql = sum(monthly[m]['lost'] for m in qm)
            qa = sum(monthly[m]['members'] for m in qm) / len(qm)
            vals = [f"{yr} Q{q}", monthly[qm[0]]['members'], qn, ql,
                    monthly[qm[-1]]['members'], qn - ql, ql / qa if qa else 0]
            for c, v in enumerate(vals, 1):
                cell = ws4.cell(row=row, column=c, value=v)
                cell.border = thin
                if c == 7: cell.number_format = pct
            row += 1
    auto_width(ws4)

    # 메모리에 저장
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ──────────────────── Streamlit UI ────────────────────

st.set_page_config(page_title="국민연금 입퇴사 분석", page_icon="📊", layout="centered")

st.markdown("""
<style>
    .main-title { text-align: center; color: #1B3A5C; margin-bottom: 0; }
    .sub-title { text-align: center; color: #888; font-size: 0.95em; margin-top: 0; }
    .stDataFrame { font-size: 0.85em; }
</style>
""", unsafe_allow_html=True)

st.markdown("<h1 class='main-title'>국민연금 입퇴사 분석</h1>", unsafe_allow_html=True)
st.markdown("<p class='sub-title'>사업장명을 검색하고, 분석 기간을 선택하세요</p>", unsafe_allow_html=True)
st.divider()

# Session state 초기화
if 'search_results' not in st.session_state:
    st.session_state.search_results = []
if 'selected_company' not in st.session_state:
    st.session_state.selected_company = None

# ── 1. 사업장 검색 ──
st.subheader("1. 사업장 검색")
col1, col2 = st.columns([4, 1])
with col1:
    keyword = st.text_input("사업장명", placeholder="예: 삼성전자, 현대자동차, 파주전기초자 ...",
                            label_visibility="collapsed")
with col2:
    search_clicked = st.button("🔍 검색", use_container_width=True)

if search_clicked and keyword.strip():
    with st.spinner("검색 중..."):
        results = search_companies(keyword.strip())
        st.session_state.search_results = results
        st.session_state.selected_company = None

# ── 2. 사업장 선택 ──
if st.session_state.search_results:
    st.subheader("2. 사업장 선택")
    results = st.session_state.search_results
    st.caption(f"총 {len(results)}건 조회됨")

    # 라디오 버튼으로 선택
    options = []
    for item in results:
        members_str = f"{item['members']:,}명" if item['members'] else '-'
        label = f"**{item['name']}** — {members_str} | {item['biz_type']} | {item['address'][:30] if item['address'] else ''}"
        options.append(label)

    selected_idx = st.radio(
        "사업장을 선택하세요",
        range(len(options)),
        format_func=lambda i: options[i],
        label_visibility="collapsed",
    )
    st.session_state.selected_company = results[selected_idx]['name']
    st.success(f"✅ 선택: **{st.session_state.selected_company}**")

# ── 3. 분석 기간 & 실행 ──
if st.session_state.selected_company:
    st.subheader("3. 분석 기간 설정 & 실행")
    col_a, col_b, col_c = st.columns([2, 1, 2])
    with col_a:
        start_year = st.selectbox("조회 시작년도", list(range(2015, 2027)), index=5)
    with col_b:
        st.markdown("<br>", unsafe_allow_html=True)
        st.write("~ 최신 데이터")

    run_clicked = st.button("📊 분석 시작", type="primary", use_container_width=True)

    if run_clicked:
        company = st.session_state.selected_company
        progress_bar = st.progress(0, text="분석 준비 중...")
        status = st.empty()

        excel_buf = run_analysis(company, start_year, progress_bar, status)

        if excel_buf:
            progress_bar.progress(1.0, text="완료!")
            st.balloons()

            safe_name = company.replace('/', '_').replace('\\', '_')
            filename = f"{safe_name}_입퇴사분석.xlsx"

            st.download_button(
                label=f"📥 {filename} 다운로드",
                data=excel_buf,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )

# ── 하단 안내 ──
st.divider()
st.caption("데이터 출처: 국민연금 가입 사업장 내역 공공데이터 (api.odcloud.kr) | 일부 누락 월은 보간 처리됩니다")
